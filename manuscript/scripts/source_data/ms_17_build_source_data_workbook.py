#!/usr/bin/env python3
"""Build figure-specific Source Data Excel workbooks with one sheet per panel."""

from __future__ import annotations

import csv
import re
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[3]
MANIFEST = ROOT / "manuscript" / "submission_data" / "source_data" / "final_source_data_manifest.tsv"
PANEL_MANIFEST = ROOT / "manuscript" / "submission_data" / "source_data" / "final_source_data_panel_manifest.tsv"
FILE_MANIFEST = ROOT / "manuscript" / "submission_data" / "source_data" / "final_source_data_file_manifest.tsv"
OUT_DIR = ROOT / "manuscript" / "submission_data" / "source_data"
OUT_README = OUT_DIR / "README.md"


def expand_panel_label(label: str) -> list[str]:
    label = label.strip()
    if re.fullmatch(r"[A-Z]-[A-Z]", label):
        start, end = label.split("-")
        return [chr(code) for code in range(ord(start), ord(end) + 1)]
    return [label]


def workbook_basename(figure_label: str) -> str:
    if figure_label.startswith("Supplementary Figure"):
        number = re.search(r"Supplementary Figure (\d+)", figure_label).group(1)
        return f"Source_Data_Supplementary_Figure_{number}.xlsx"
    number = re.search(r"Figure (\d+)", figure_label).group(1)
    return f"Source_Data_Figure_{number}.xlsx"


def figure_sort_key(figure_label: str) -> tuple[int, int]:
    if figure_label.startswith("Supplementary Figure"):
        number = int(re.search(r"Supplementary Figure (\d+)", figure_label).group(1))
        return (1, number)
    number = int(re.search(r"Figure (\d+)", figure_label).group(1))
    return (0, number)


def read_manifest() -> list[dict[str, str]]:
    with MANIFEST.open("r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        return list(reader)


def build_panel_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    expanded: list[dict[str, str]] = []
    for row in rows:
        panels = expand_panel_label(row["panel_label"])
        grouped = len(panels) > 1
        workbook_file = workbook_basename(row["figure_label"])
        for panel in panels:
            expanded.append(
                {
                    "figure_label": row["figure_label"],
                    "panel_label": panel,
                    "panel_title": row["panel_title"],
                    "final_placement": row["final_placement"],
                    "frozen_input_files": row["frozen_input_files"],
                    "source_bundle_shared_across_panels": "yes" if grouped else "no",
                    "source_bundle_original_panel_label": row["panel_label"],
                    "sheet_name": f"Panel_{panel}",
                    "workbook_file": workbook_file,
                    "workbook_relative_path": (OUT_DIR / workbook_file).relative_to(ROOT).as_posix(),
                }
            )
    return expanded


def write_panel_manifest(rows: list[dict[str, str]]) -> None:
    PANEL_MANIFEST.parent.mkdir(parents=True, exist_ok=True)
    with PANEL_MANIFEST.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "workbook_file",
                "workbook_relative_path",
                "sheet_name",
                "figure_label",
                "panel_label",
                "panel_title",
                "final_placement",
                "source_bundle_shared_across_panels",
                "source_bundle_original_panel_label",
                "frozen_input_files",
            ],
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_file_manifest(rows: list[dict[str, str]]) -> None:
    grouped: dict[str, dict[str, str]] = {}
    for row in rows:
        figure_label = row["figure_label"]
        if figure_label not in grouped:
            grouped[figure_label] = {
                "figure_label": figure_label,
                "final_placement": row["final_placement"],
                "workbook_file": row["workbook_file"],
                "workbook_relative_path": row["workbook_relative_path"],
                "panel_labels": row["panel_label"],
                "panel_count": "1",
            }
        else:
            grouped[figure_label]["panel_labels"] += f";{row['panel_label']}"
            grouped[figure_label]["panel_count"] = str(int(grouped[figure_label]["panel_count"]) + 1)

    ordered = [grouped[label] for label in sorted(grouped, key=figure_sort_key)]
    with FILE_MANIFEST.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "figure_label",
                "final_placement",
                "workbook_file",
                "workbook_relative_path",
                "panel_count",
                "panel_labels",
            ],
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(ordered)


def load_table(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", dtype=str, keep_default_na=False)


def autosize_sheet(ws) -> None:
    max_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            width = len(str(cell.value))
            max_widths[cell.column] = min(max(max_widths.get(cell.column, 0), width), 60)
    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def build_workbooks(rows: list[dict[str, str]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for old_file in OUT_DIR.glob("Source_Data*.xlsx"):
        old_file.unlink()

    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        grouped.setdefault(row["figure_label"], []).append(row)

    for figure_label in sorted(grouped, key=figure_sort_key):
        figure_rows = sorted(grouped[figure_label], key=lambda item: item["panel_label"])
        workbook_path = OUT_DIR / figure_rows[0]["workbook_file"]
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            for row in figure_rows:
                sheet_name = row["sheet_name"][:31]
                meta_df = pd.DataFrame(
                    [
                        ["figure_label", row["figure_label"]],
                        ["panel_label", row["panel_label"]],
                        ["panel_title", row["panel_title"]],
                        ["final_placement", row["final_placement"]],
                        ["shared_input_bundle", row["source_bundle_shared_across_panels"]],
                        ["original_bundle_label", row["source_bundle_original_panel_label"]],
                    ],
                    columns=["field", "value"],
                )
                meta_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                startrow = len(meta_df) + 2

                for source_file in [item.strip() for item in row["frozen_input_files"].split(";") if item.strip()]:
                    source_path = ROOT / source_file
                    header_df = pd.DataFrame(
                        [
                            {
                                "source_file": source_file,
                                "status": "found" if source_path.exists() else "missing",
                            }
                        ]
                    )
                    header_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
                    startrow += len(header_df) + 2
                    if source_path.exists():
                        table_df = load_table(source_path)
                        table_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
                        startrow += len(table_df.index) + 3
                    else:
                        startrow += 2

            for name in writer.book.sheetnames:
                ws = writer.book[name]
                ws.freeze_panes = "A2"
                autosize_sheet(ws)


def write_readme(rows: list[dict[str, str]]) -> None:
    grouped: dict[str, dict[str, str]] = {}
    for row in rows:
        if row["figure_label"] not in grouped:
            grouped[row["figure_label"]] = {
                "workbook_file": row["workbook_file"],
                "workbook_relative_path": row["workbook_relative_path"],
                "final_placement": row["final_placement"],
            }

    main_lines = []
    extended_lines = []
    for figure_label in sorted(grouped, key=figure_sort_key):
        info = grouped[figure_label]
        target = info["workbook_file"]
        label = figure_label if figure_label.startswith("Supplementary Figure") else info["workbook_file"]
        line = f"- [{label}]({target})"
        if info["final_placement"] == "main":
            main_lines.append(line)
        else:
            extended_lines.append(line)

    content = "\n".join(
        [
            "# Source Data",
            "",
            "This directory contains one Excel workbook per figure for the submission package.",
            "Each workbook contains one sheet per panel, with source-file headers followed by the exported table blocks used to build that panel.",
            "",
            f"`Base manifest`: `{MANIFEST.relative_to(ROOT).as_posix()}`",
            f"`Panel manifest`: `{PANEL_MANIFEST.relative_to(ROOT).as_posix()}`",
            f"`File manifest`: `{FILE_MANIFEST.relative_to(ROOT).as_posix()}`",
            "",
            "## Main Figures",
            *main_lines,
            "",
            "## Supplementary Figures",
            *extended_lines,
            "",
        ]
    )
    OUT_README.write_text(content + "\n", encoding="utf-8")


def main() -> None:
    base_rows = read_manifest()
    panel_rows = build_panel_rows(base_rows)
    write_panel_manifest(panel_rows)
    write_file_manifest(panel_rows)
    build_workbooks(panel_rows)
    write_readme(panel_rows)

if __name__ == "__main__":
    main()
